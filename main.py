import sys
import os
import shutil
import pandas as pd
import time
from pathlib import Path
from datetime import datetime
from playwright.sync_api import Playwright, sync_playwright
from openpyxl.styles import Border, Side, Alignment
from dotenv import load_dotenv

import gmail
from send_error import send_message_to_shell

load_dotenv()

PROJECT_PATH = os.getenv("PROJECT_PATH")

def cjoy_login(page):
    """로그인 함수"""
    try:
        site_url = os.getenv("SITE_URL")
        email = os.getenv("EMAIL")
        password = os.getenv("PASSWORD")

        # page.goto(site_url)
        page.goto(site_url, wait_until="domcontentloaded", timeout=60000)
        page.get_by_test_id("email").click()
        page.get_by_test_id("email").fill(email)
        page.get_by_test_id("password").click()
        page.get_by_test_id("password").fill(password)
        page.get_by_test_id("submit").click()
    except Exception as e:
        send_message_to_shell(str(e))
        
def download_campaign_report(page, period):
    """캠페인 리포트 다운로드 함수 (period: 'thisMonth' 또는 'lastMonth')"""
    
    print(f"*************캠페인 리포트 다운로드 ({period})*************")
    try:
        page.wait_for_load_state("domcontentloaded", timeout=60000)
        page.get_by_test_id("homeDateRangePickerReference").click()
        
        if period == "thisMonth":
            page.get_by_test_id("thisMonth").click()
        elif period == "lastMonth":
            page.get_by_test_id("lastMonth").click()
        
        page.get_by_test_id("homeDateRangePickerDialogApply").click()
        # page.get_by_label("다운로드").click()

        # 파일 다운로드
        with page.expect_download() as download_info:
            page.get_by_label("캠페인 리포트").click()
        download = download_info.value
            
        # 다운로드 파일 경로 얻기 (Path 객체로 변환)
        download_path = Path(download.path())
        print(f"{period} 다운로드한 파일 경로: {download_path}")

        # 프로젝트 다운로드 경로
        PROJECT_DOWNLOADS_PATH = os.path.join(PROJECT_PATH, "downloads")
        target_download_path = ''
        # .csv 파일 경로 설정
        if not str(download_path).endswith('.csv'):
            new_download_path = download_path.with_suffix('.csv')  # .csv 확장자 추가
            
            # 다운로드한 파일의 새로운 경로 설정
            target_download_path = os.path.join(PROJECT_DOWNLOADS_PATH, new_download_path.name)

            # 파일 이동
            shutil.move(str(download_path), str(target_download_path))  # 파일 이동
            print(f"파일이 {target_download_path}로 이동되었습니다.")
        else:
            target_download_path = os.path.join(PROJECT_DOWNLOADS_PATH, download_path.name)
            
            # 파일이 이미 csv 확장자인 경우 이동
            shutil.move(str(download_path), str(target_download_path))
            print(f"파일이 {target_download_path}로 이동되었습니다.")
            
    except Exception as e:
        send_message_to_shell(str(e))
        
    return target_download_path  # 이동한 파일 경로 반환

def save_csv_as_excel(download_path, period):
    """CSV 파일을 엑셀 파일로 변환하고 저장하는 함수"""
    print("########csv를 엑셀로 변환합니다########")
    try:
        # 이후 CSV 파일 처리
        df = pd.read_csv(download_path)

        # '집행 광고비'가 '0'인 행 삭제
        df = df[df['집행 광고비'] != '0']


        # 엑셀 파일 이름 설정
        if period == "thisMonth" or period == "lastMonth":
            # 해당 기간에 맞는 열 순서
            column_order = [
                "캠페인", "일 예산", "입찰가(CPC)", "집행 광고비", "노출 수",
                "클릭 수", "클릭당 광고비", "클릭률", "CPM", "판매 수",
                "직접 전환 판매 수", "간접 전환 판매 수", "전환 매출",
                "직접 전환 매출", "간접 전환 매출", "광고 수익률(ROAS)",
                "직접 광고 수익률(ROAS)", "간접 광고 수익률(ROAS)", "전환율"
            ]
        else:
            # 기본 열 순서
            column_order = [
                "상품", "집행 광고비", "노출 수",
                "클릭 수", "클릭당 광고비", "클릭률", "CPM", "판매 수",
                "직접 전환 판매 수", "간접 전환 판매 수", "전환 매출",
                "직접 전환 매출", "간접 전환 매출", "광고 수익률(ROAS)",
                "직접 광고 수익률(ROAS)", "간접 광고 수익률(ROAS)", "전환율"
            ]
        
        # 지정된 열만 포함하고 순서대로 정렬
        df = df[column_order]  # 지정된 순서의 열만 포함


        # 엑셀 파일 이름 설정
        if period == "thisMonth":
            output_filename = f"더샘_성과형광고_토탈효율_저장_{datetime.now().strftime('%Y%m%d')}.xlsx"
        elif period == "lastMonth":
            output_filename = f"더샘_성과형광고_토탈효율_저장_전월.xlsx"
        else:
            output_filename = f"더샘_성과형광고_{period}_효율_저장_{datetime.now().strftime('%Y%m%d')}.xlsx"


        # 저장할 경로 설정
        save_directory = os.path.join(PROJECT_PATH, "excel")

        output_path = os.path.join(save_directory, output_filename)

        # 이미 존재하는 파일이 있을 경우 삭제
        if os.path.exists(output_path):
            os.remove(output_path)
            print(f"기존 파일 {output_path}를 삭제했습니다.")

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 인덱스 열 추가
            df.insert(0, ' ', range(1, len(df) + 1))  # A열 인덱스를 1부터 시작하도록 추가
            
            df.to_excel(writer, index=False, startrow=0, header=True)  # 제목 행을 포함하여 Excel 파일에 쓰기

            # 현재 시트에서 워크북 가져오기
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # A1 셀 값을 빈 문자열로 설정
            worksheet['A1'] = ''

            # 1행 텍스트 줄 바꿈 및 중앙 정렬
            for cell in worksheet[1]:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')  # 줄 바꿈 및 중앙 정렬

            # 두 번째 열 높이 15로 수정
            worksheet.row_dimensions[2].height = 15

            # B행의 열 너비 30으로 수정
            worksheet.column_dimensions['B'].width = 30
            # C행의 열 너비 12로 수정
            worksheet.column_dimensions['C'].width = 12

            # C2 셀의 소수점 없애기
            worksheet['C2'].number_format = '0'  # 소수점 없이 표시

            # 특정 셀에 테두리 추가
            thin = Side(style='thin')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            worksheet['C2'].border = border  # C2 셀에 테두리 추가

            # 전체 데이터 범위에 테두리 그리기
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        print(f"파일이 {output_path}에 저장되었습니다.")
    except Exception as e:
        send_message_to_shell(str(e))

def make_excel_for_performance_ad_campaign_product_efficiency(page):
    """성과형 광고 캠페인 상품 효율 리포트 다운로드 및 엑셀로 변환하는 함수"""
    
    print("*************성과형 광고 캠페인 별 제품 효율 (당월)*************")
    try:
        page.get_by_role("link", name="광고 홈").click()
        page.get_by_test_id("adAccountDashboardTableContainer").get_by_role("combobox").select_option("50")
        page.get_by_test_id("homeDateRangePickerReference").click()
        page.get_by_test_id("thisMonth").click()
        page.get_by_test_id("homeDateRangePickerDialogApply").click()

        # aria-rowindex 속성을 가진 모든 div 요소 선택
        rows_with_aria = page.locator('div[aria-rowindex]')  # aria-rowindex 속성이 있는 div 선택

        # 데이터 저장을 위한 리스트
        data = []

        # aria-rowindex 속성을 가진 각 행에서 데이터 추출
        # 캠페인 그리드 파싱
        count = rows_with_aria.count()  # 비동기 메서드 호출

        for i in range(count):  # count를 사용하여 반복
            row = rows_with_aria.nth(i)
            aria_rowindex_value = row.get_attribute('aria-rowindex')  # aria-rowindex 속성 값 가져오기
            row_text = row.inner_text()  # 행의 텍스트 내용 가져오기
            data.append({
                'aria_rowindex': aria_rowindex_value,
                'text': row_text
            })

        # 파싱된 캠페인 그리드 DataFrame으로 변환

        # 첫 번째 항목(타이틀 행)을 기준으로 열 이름 생성
        columns = data[0]['text'].split('\n')

        # 각 행의 데이터를 split하여 리스트로 저장
        rows = [item['text'].split('\n') for item in data[1:]]

        # DataFrame 생성 전에 행의 개수 확인 후 부족한 열에 빈 값(None) 추가
        for row in rows:
            while len(row) < len(columns):
                row.append(None)

        # pandas DataFrame 생성
        df = pd.DataFrame(rows, columns=columns)

        # 테이블 출력
        print(df)
        
        state = '활성'
        
        # 상태가 '활성'인 캠페인의 캠페인명만 필터링하여 리스트로 변환
        active_campaign_names = df[df['상태'] == state]['캠페인'].tolist()
        print('#' * 50)
        print(f"{state} 상태인 캠페인")
        print(active_campaign_names)
        print('#' * 50)
        if not active_campaign_names:
            print(f"{state} 상태인 캠페인이 없습니다.")
            return
        # 각 캠페인 이름에 해당하는 <div> 요소 클릭
        for campaign_name in active_campaign_names:
            # 캠페인 요소를 찾기 위해 광고홈 이동
            print('캠페인 요소를 찾기 위해 광고홈 이동')
            page.get_by_role("link", name="광고 홈").click()
            page.get_by_test_id("adAccountDashboardTableContainer").get_by_role("combobox").select_option("50")
            time.sleep(1)

            # adAccountDashboardTable 컨테이너 내에서 캠페인 이름을 찾기 위해 반복 스크롤
            found = False
            
            for _ in range(10):  # 최대 10회 스크롤 반복
                print("adAccountDashboardTable 컨테이너 아래로 스크롤 중...")

                # 컨테이너 직접 선택하고 스크롤
                page.evaluate("document.querySelector('[data-testid=\"adAccountDashboardTable\"]').scrollBy(0, 500);")
                time.sleep(1)  # 스크롤 후 대기

                # 캠페인 이름이 포함된 요소 찾기 (첫 번째 요소 선택)
                locator = page.locator(f"[data-testid='adAccountDashboardTable'] div:has-text('{campaign_name}')").nth(1)
                
                if locator.is_visible():
                    locator.click()
                    print(f"캠페인 '{campaign_name}'을(를) 클릭했습니다.")
                    found = True
                    break
                else:
                    print(f"캠페인 '{campaign_name}'을 찾지 못해 스크롤을 계속합니다.")

            # 요소를 찾지 못한 경우 메시지 출력
            if not found:
                error_message = f"캠페인 '{campaign_name}'을 찾을 수 없습니다."
                print(error_message)
                send_message_to_shell(error_message)

            # 캠페인 상품 리포트 다운로드을 위해 '상품'탭 클릭
            page.get_by_label("item").click()

            # 비동기 컨텍스트 매니저 사용
            with page.expect_download() as download_info:
                page.get_by_label("상품 리포트").click()
            download = download_info.value

            # 다운로드 파일 경로 얻기 (await download.path() 사용)
            download_path = download.path()  # await 추가
            print('#' * 50)
            print(f"캠페인[{campaign_name}] 다운로드한 csv 파일 경로: {download_path}")
            print('#' * 50)
            save_csv_as_excel(download_path, campaign_name)
    except Exception as e:
        send_message_to_shell(str(e))

def ensure_directory_exists(directory):
    """주어진 디렉토리가 존재하지 않으면 생성합니다."""
    if not os.path.exists(directory):
        os.makedirs(directory)
        print(f"디렉토리 '{directory}'가 생성되었습니다.")

def get_files_from_directory(directory, extension=None):
    """
    특정 디렉토리에서 모든 파일을 가져오거나 특정 확장자의 파일만 가져옴.
    파일이 없는 경우 빈 리스트를 반환함.
    
    :param directory: 검색할 디렉토리 경로
    :param extension: 특정 파일 확장자 (예: '.txt'). None일 경우 모든 파일을 반환.
    :return: 해당 디렉토리의 파일 경로 리스트. 파일이 없으면 빈 리스트.
    """
    files = []
    
    # 디렉토리 내 파일 목록 확인
    for file in os.listdir(directory):
        file_path = os.path.join(directory, file)
        
        # 파일이면서 확장자가 조건에 맞는 경우 리스트에 추가
        if os.path.isfile(file_path) and (extension is None or file.endswith(extension)):
            files.append(file_path)
    
    # 파일이 없는 경우 빈 리스트 반환
    return files  
        
def setup_directories(project_path):
    """
    필요한 폴더들을 생성하고, 기준일자 폴더 내 파일을 확인합니다.
    기준일자 폴더에 파일이 있으면 프로그램을 종료합니다.
    
    :param project_path: 프로젝트 경로
    """
    # 전송 폴더 및 기준일자 폴더 생성 및 파일 확인
    send_directory = os.path.join(project_path, "send")
    date_folder_name = datetime.now().strftime('%Y%m%d')
    send_date_folder_path = os.path.join(send_directory, date_folder_name)

    # 기준일자 폴더 확인 및 생성
    ensure_directory_exists(send_date_folder_path)

    # 기준일자 폴더에 이미 전송된 파일이 있는지 확인
    send_files = get_files_from_directory(send_date_folder_path, extension=".xlsx")
    if send_files:
        print(f"{datetime.now().strftime('%Y%m%d')} 이미 이메일이 발송되어 종료합니다.")
        sys.exit(0)

    # 다운로드, 엑셀 폴더 경로 생성
    downloads_folder = os.path.join(project_path, "downloads")
    excel_folder = os.path.join(project_path, "excel")
    send_folder = os.path.join(project_path, "send")

    # 폴더들 확인 및 생성
    ensure_directory_exists(downloads_folder)
    ensure_directory_exists(excel_folder)
    ensure_directory_exists(send_folder)

def setup_browser(playwright: Playwright):
    """Playwright 브라우저 설정을 처리합니다."""
    env = os.getenv('ENV')
    print(f"현재 환경: {env}")
    
    headless = False
    if env == 'production':
        headless = True

    # 사용자 데이터 디렉토리를 명시적으로 지정
    user_data_dir = os.path.join(PROJECT_PATH, "playwright_cache")
    
    # 캐시 비활성화 및 사용자 데이터 디렉토리 설정
    # browser = playwright.chromium.launch(
    #     headless=headless,
    #     args=["--disable-cache"],
    #     user_data_dir=user_data_dir
    # )
    # context = browser.new_context(
    #     locale="ko-KR", bypass_csp=True, ignore_https_errors=True
    # )
    # 캐시 비활성화 및 사용자 데이터 디렉토리 설정
    context = playwright.chromium.launch_persistent_context(
        user_data_dir=user_data_dir,  # 사용자 데이터 디렉토리 설정
        headless=headless,            # headless 모드 설정
        args=["--disable-cache"],     # 캐시 비활성화
        locale="ko-KR",               # 로케일 설정
        bypass_csp=True,              # CSP 무시
        ignore_https_errors=True      # HTTPS 인증서 오류 무시
    )
    page = context.new_page()

    return context, page



def run(playwright: Playwright) -> None:
    try:
        # 폴더 생성 및 파일 확인
        setup_directories(PROJECT_PATH)    
        
        # 브라우저 설정 및 실행
        context, page = setup_browser(playwright)
        
        # 로그인
        cjoy_login(page)

        # 캠페인 리포트 다운로드 (당월)
        this_month_download_path = download_campaign_report(page, "thisMonth")
        save_csv_as_excel(this_month_download_path, "thisMonth")

        # 캠페인 리포트 다운로드 (전월)
        last_month_download_path = download_campaign_report(page, "lastMonth")
        save_csv_as_excel(last_month_download_path, "lastMonth")

        # 성과형 광고 캠페인 별 제품 효율 (당월)
        make_excel_for_performance_ad_campaign_product_efficiency(page)
        
        # # ---------------------
        context.close()

        # 이메일 전송
        gmail.main()
    except Exception as e:
        send_message_to_shell(str(e))
        
with sync_playwright() as playwright:
    run(playwright)
