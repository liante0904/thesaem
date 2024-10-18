import requests
import json
import time
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# 키워드 리스트
keywords = ["롬앤", "더모아", "아이폰"]

# 현재 시간을 밀리초로 변환
current_time = int(time.time() * 1000)

# 결과를 저장할 리스트
data = {
    'Keyword': [],
    'POST_result_1': [],
    'POST_result_2': [],
    'POST_result_3': [],
    'POST_result_4': [],
    'POST_result_5': [],
    'POST_result_6': [],
    'POST_result_7': [],
    'POST_result_8': [],
    'GET_shopCategory': [],
    'GET_monthBlog': [],
    'GET_blogSaturation': []
}

# 헤더 설정
headers = {
    "referer": "https://www.ma-pia.net/",
    "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36"
}

# POST와 GET 요청을 통해 데이터를 수집
for keyword in keywords:
    print(f"현재 키워드: {keyword}")
    
    # POST 요청
    post_url = "https://www.ma-pia.net/keyword/keywordQ.php"
    post_data = {'DataQ': keyword}
    
    post_response = requests.post(post_url, data=post_data, headers=headers)
    if post_response.status_code == 200:
        print(f"POST 요청 성공: {keyword}")
        print(post_response.text)
        post_result = post_response.text.split("success :")[1].split("|||")[0].split("///")
        
        # POST 결과 추가
        data['Keyword'].append(keyword)
        data['POST_result_1'].append(post_result[1])
        data['POST_result_2'].append(post_result[2])
        data['POST_result_3'].append(post_result[3])
        data['POST_result_4'].append(post_result[4])
        data['POST_result_5'].append(post_result[5])
        data['POST_result_6'].append(post_result[6])
        data['POST_result_7'].append(post_result[7])
        data['POST_result_8'].append(post_result[8])
    else:
        print(f"POST 요청 실패: {keyword}")
        continue
    
    # GET 요청
    get_url = "https://uy3w6h3mzi.execute-api.ap-northeast-2.amazonaws.com/Prod/hello"
    get_params = {
        "keyword": keyword,
        "totalSum": int(post_result[1])+int(post_result[2]),
        "time": current_time
    }
    get_response = requests.get(get_url, params=get_params, headers=headers)
    
    if get_response.status_code == 200:
        print(f"GET 요청 성공: {keyword}")
        get_data = get_response.json()
        
        # GET 결과 추가
        data['GET_shopCategory'].append(get_data['result']['shopCategory'])
        data['GET_monthBlog'].append(get_data['result']['monthBlog'])
        data['GET_blogSaturation'].append(get_data['result']['blogSaturation'])
    else:
        print(f"GET 요청 실패: {keyword}")
        continue

# DataFrame 생성
df = pd.DataFrame(data)
print(data)

# 새로운 엑셀 파일 생성
wb = Workbook()
ws = wb.active




# 첫 번째 행에 안내 문구 추가(공백셀)
ws.merge_cells('A1:M1')
ws['A1'] = ' '

# 두 번째 행에 컬럼 이름 추가
columns = [
    "키워드", "월간검색수", "", "검색수\n합계", "월간 블로그 발행", "", 
    "네이버쇼핑\n카테고리", "월평균클릭수", "", "월평균클릭율(%)", "", "경쟁정도", "월평균\n노출광고수"
]
ws.append(columns)

# 셀 병합 처리
ws.merge_cells('B2:C2')
ws.merge_cells('H2:I2')
ws.merge_cells('J2:K2')

# 추가 병합
ws.merge_cells('A2:A3')  # A2:A3 병합
ws.merge_cells('D2:D3')  # D2:D3 병합
ws.merge_cells('E2:F2')  # E2:F2 병합
ws.merge_cells('G2:G3')  # G2:G3 병합
ws.merge_cells('L2:L3')  # L2:L3 병합
ws.merge_cells('M2:M3')  # M2:M3 병합

# 각 병합된 셀에 값을 설정
ws['A2'] = '키워드'  # A2 병합된 셀 값
ws['D2'] = '검색수\n합계'  # D2 병합된 셀 값
ws['E2'] = '월간 블로그 발행'  # E2 병합된 셀 값
ws['G2'] = '네이버쇼핑\n카테고리'  # G2 병합된 셀 값
ws['L2'] = '경쟁정도'  # L2 병합된 셀 값
ws['M2'] = '월평균\n노출광고'  # M2 병합된 셀 값

# 세 번째 행에 서브 컬럼 이름 추가
sub_columns = [
    "", "PC", "모바일", "", "수량", "포화도", "", "PC", "모바일", "PC", "모바일", "", ""
]
ws.append(sub_columns)

# 데이터 추가
for i in range(len(data['Keyword'])):
    row = [
        data['Keyword'][i],
        data['POST_result_1'][i],
        data['POST_result_2'][i],
        int(data['POST_result_1'][i])+int(data['POST_result_2'][i]),
        data['GET_monthBlog'][i],
        data['GET_blogSaturation'][i],
        data['GET_shopCategory'][i],
        data['POST_result_3'][i],
        data['POST_result_4'][i],
        data['POST_result_5'][i],
        data['POST_result_6'][i],
        data['POST_result_7'][i],
        data['POST_result_8'][i]
        
    ]
    ws.append(row)


# 전체 셀에 대해 맑은 고딕 글꼴 적용
malgun_gothic_font = Font(name='맑은 고딕')

for row in ws.iter_rows():
    for cell in row:
        cell.font = malgun_gothic_font


# 정렬 설정 (병합된 셀도 고려)
for merged_range in ws.merged_cells.ranges:
    min_col, min_row, max_col, max_row = merged_range.bounds
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')


# 각 열의 셀 정렬 설정
for row in range(1, len(data['Keyword']) + 4):  # 데이터가 있는 행까지
    for col in [2, 3]:  # 2열과 3열에 대해서만
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')

# G열과 L열의 텍스트 가운데 정렬
for row in range(1, len(data['Keyword']) + 4):  # 데이터가 있는 행까지
    for col in [7, 12]:  # G열(7)과 L열(12)
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')

# B~D열의 4열부터 유효 데이터가 있는 셀 우측 정렬 및 천 단위 구분 기호 추가
for row in range(4, len(data['Keyword']) + 4):  # 4열부터 시작
    for col in range(2, 5):  # B열(2)부터 D열(4)까지
        cell_value = ws.cell(row=row, column=col).value
        if cell_value is not None:  # 유효 데이터가 있는 셀만
            # 숫자형 데이터인지 확인
            if isinstance(cell_value, (int, float)) or (isinstance(cell_value, str) and cell_value.isdigit() and int(cell_value) > 999):
                if isinstance(cell_value, str):
                    cell_value = int(cell_value)  # 문자열을 정수로 변환
                ws.cell(row=row, column=col).value = f"{cell_value:,}"  # 천 단위 구분 기호 추가
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='right', vertical='center')

# E열과 F열 우측 정렬 추가
for row in range(4, len(data['Keyword']) + 4):  # 4열부터 시작
    for col in range(5, 7):  # E열(5)과 F열(6)
        cell_value = ws.cell(row=row, column=col).value
        if cell_value is not None:  # 유효 데이터가 있는 셀만
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='right', vertical='center')


# H~K열과 M열의 4열부터 유효 데이터가 있는 셀 우측 정렬
for row in range(4, len(data['Keyword']) + 4):  # 4열부터 시작
    for col in range(8, 12):  # H열(8)부터 K열(11)까지
        cell_value = ws.cell(row=row, column=col).value
        if cell_value is not None:  # 유효 데이터가 있는 셀만
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='right', vertical='center')
    # M열(13)도 추가
    cell_value = ws.cell(row=row, column=13).value
    if cell_value is not None:  # 유효 데이터가 있는 셀만
        ws.cell(row=row, column=13).alignment = Alignment(horizontal='right', vertical='center')

# 병합된 셀에서 \n 문자가 있는 경우 '자동 줄바꿈'과 '가운데 정렬' 설정
for merged_range in ws.merged_cells.ranges:
    for row in ws[merged_range.coord]:
        for cell in row:
            cell_value = cell.value
            if cell_value and isinstance(cell_value, str) and '\n' in cell_value:
                # 자동 줄바꿈 및 가운데 정렬 활성화
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# E~F, H~K 열의 3열 가운데 정렬 설정
for col in range(5, 7):  # E열(5)과 F열(6)
    ws.cell(row=3, column=col).alignment = Alignment(horizontal='center', vertical='center')

for col in range(8, 12):  # H열(8)부터 K열(11)까지
    ws.cell(row=3, column=col).alignment = Alignment(horizontal='center', vertical='center')


# 모든 열 너비 설정 (G열 제외)
for col_idx in range(1, ws.max_column + 1):  # 1부터 최대 열까지 반복
    col_letter = get_column_letter(col_idx)  # 열 문자를 열 인덱스로부터 가져옴
    if col_letter == 'G':
        ws.column_dimensions[col_letter].width = 14  # G열 너비는 14로 설정
    else:
        ws.column_dimensions[col_letter].width = 9  # 나머지 열은 8로 설정


# 파일 저장
output_path = 'mapianet_results.xlsx'
wb.save(output_path)
print("엑셀 파일 생성 완료.")